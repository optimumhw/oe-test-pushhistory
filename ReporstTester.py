__author__ = 'hal'

import datetime
import pymssql
import requests
import json
import urllib
import os
import openpyxl

class ConfigurationHelper:
    def __init__(self):

        try:
            self.configDirPath = os.environ['EDGE_PYTHON_SCRIPTS_CONFIG_HOME'].strip()
        except:
            print

            #self.configDirPath = '/Users/halwilkinson/Desktop/EdgePythonScriptsConfig'  #home
            self.configDirPath = '/Users/hal/Desktop/PythonScriptConfigDir' #work
            print 'could not read EDGE_PYTHON_SCRIPTS_CONFIG_HOME env variable. using:' + self.configDirPath
            pass

        try:
            configFilePath = os.path.join(self.configDirPath, 'config.txt')
            configFile = open(configFilePath, 'r')
            configLines = configFile.read().splitlines()
            configFile.close()

            self.configSettings = {}
            for line in configLines:
                if len(line) <= 0:
                    continue
                keyAndValue = line.split('=')
                self.configSettings[keyAndValue[0]] = keyAndValue[1]

        except:
            logError('config.txt file i/o error')

    def getConfigSettings(self):
        return self.configSettings

    def getConfigDirPath(self):
        return self.configDirPath


class ExcelHelper():
    def __init__(self, configDirPath, reportsDataFileName):

        self.e3osResultsSheetName = 'E3OSPtHist'
        self.edgeResultsSheetName = 'EdgePtHist'
        self.dependentPointHistorSheetName = 'DepPtHist'
        self.calcResultsSheetName = 'CalcPtHist'

        self.spreadSheetName = reportsDataFileName

        excelFilePath = os.path.join(configDirPath, reportsDataFileName)

        self.resultsFilePath = os.path.join(configDirPath, 'foobar.xlsx')
        self.workbook = openpyxl.load_workbook(excelFilePath)
        #self.pointNameSheet = self.workbook.get_sheet_by_name('Points')
        #self.dependentPointsSheet = self.workbook.get_sheet_by_name('DependentPoints')
        #self.calculatedPointsSheet = self.workbook.get_sheet_by_name('CalculatedPoints')


    def getExcelPointsAndTypes(self, sheetName):

        sheet = self.workbook.get_sheet_by_name(sheetName)

        pointsNameAndTypeList = []
        isFirstRow = True
        for rowNumber in range(0, sheet.max_row):

            if isFirstRow:
                isFirstRow = False
                continue

            includeCellObj = sheet['A' + str(rowNumber+1)]
            e3osPointNameCellObj = sheet['B' + str(rowNumber+1)]
            pointTypeCellObj = sheet['C' + str(rowNumber+1)]
            edisonPointNameCellObj = sheet['D' + str(rowNumber+1)]
            aggregateTypeCellObj  = sheet['E' + str(rowNumber+1)]

            includeFlag = includeCellObj.value
            e3osPointName = e3osPointNameCellObj.value
            pointType = pointTypeCellObj.value
            edisonPointName = edisonPointNameCellObj.value
            aggregateTypeName = aggregateTypeCellObj.value

            if includeFlag == 0:
                continue

            dict = {}
            if e3osPointName is None or len(e3osPointName) == 0:
                dict['e3osPointName'] = edisonPointName
            else:
                dict['e3osPointName'] = e3osPointName

            dict['pointType'] = pointType
            dict['edisonPointName'] = edisonPointName
            dict['aggregateTypeName'] = aggregateTypeName

            if dict['e3osPointName'] is not None and len(dict['e3osPointName']) > 0 :
                pointsNameAndTypeList.append(dict)

        return pointsNameAndTypeList

    def mergeResults(self, returnedPointNames, e3osTimestamps, e3osTimestampsAndValues, edgeTimestamps, edgeTimestampsAndValues):

        ws = self.workbook.create_sheet(self.e3osResultsSheetName)

        ws.cell(row=1, column=1).value = 'Timestamp'
        pointIndex = 0
        for pointName in returnedPointNames:
            ws.cell(row=1, column=2 + pointIndex).value = pointName
            pointIndex = pointIndex + 1


        for tsIndex in range(0, len( e3osTimestamps )):
            ts = e3osTimestamps[tsIndex]

            ws.cell(row=tsIndex + 2, column=1).value = ts

            pointIndex = 0
            for pointName in returnedPointNames:
                values = e3osTimestampsAndValues[ts]
                e3osPointValue = values[pointIndex]
                ws.cell(row=tsIndex + 2, column=pointIndex+2).value = e3osPointValue

                pointIndex = pointIndex + 1

        self.workbook.save(self.resultsFilePath)

        ws = self.workbook.create_sheet(self.edgeResultsSheetName)

        ws.cell(row=1, column=1).value = 'Timestamp'
        pointIndex = 0
        for pointName in returnedPointNames:
            ws.cell(row=1, column=2 + pointIndex).value = pointName
            pointIndex = pointIndex + 1

        for tsIndex in range(0, len(edgeTimestamps)):
            ts = edgeTimestamps[tsIndex]

            ws.cell(row=tsIndex + 2, column=1).value = ts

            pointIndex = 0
            for pointName in returnedPointNames:
                values = edgeTimestampsAndValues[ts]
                edgePointValue = values[pointIndex]
                ws.cell(row=tsIndex + 2, column=pointIndex + 2).value = edgePointValue

                pointIndex = pointIndex + 1

        self.workbook.save(self.resultsFilePath)

    def addDependentPointHistorySheet( self, configPoints, returnedPointNames, edgeTimestamps, dependentPointsTimestampsAndValues ):

        ws = self.workbook.create_sheet(self.dependentPointHistorSheetName)

        ws.cell(row=1, column=1).value = 'config point'
        ws.cell(row=1, column=2).value = 'value'

        configPointIndex = 0
        for keyName in configPoints.keys():
            ws.cell(row=configPointIndex + 2, column=1).value = keyName
            ws.cell(row=configPointIndex + 2, column=2).value = configPoints[ keyName ]
            configPointIndex = configPointIndex + 1


        startIndex = 3 + len( configPoints.keys() )

        ws.cell(row=startIndex, column=1).value = 'Timestamp'
        pointIndex = 0
        for returnedPointName in returnedPointNames:
            ws.cell(row=startIndex, column=2+pointIndex).value = returnedPointName
            pointIndex = pointIndex + 1

        for tsIndex in range(0, len(edgeTimestamps)):
            ts = edgeTimestamps[tsIndex]
            ws.cell(row=startIndex+1+tsIndex, column=1).value = ts

            for pointIndex in range(0, len(returnedPointNames)):
                valuesAtTimestamp = dependentPointsTimestampsAndValues[ts]
                edgePointValue = valuesAtTimestamp[pointIndex]
                ws.cell(row=startIndex+1+tsIndex, column=pointIndex + 2).value = edgePointValue

        self.workbook.save(self.resultsFilePath)

    def addCalculatedPointHistorySheet( self, configPoints, returnedPointNames, edgeTimestamps, timestampsAndValues ):

        ws = self.workbook.create_sheet(self.calcResultsSheetName)

        ws.cell(row=1, column=1).value = 'config point'
        ws.cell(row=1, column=2).value = 'value'

        configPointIndex = 0
        for keyName in configPoints.keys():
            ws.cell(row=configPointIndex + 2, column=1).value = keyName
            ws.cell(row=configPointIndex + 2, column=2).value = configPoints[ keyName ]
            configPointIndex = configPointIndex + 1


        startIndex = 3 + len( configPoints.keys() )

        ws.cell(row=startIndex, column=1).value = 'Timestamp'
        pointIndex = 0
        kIndex = 0

        for returnedPointName in returnedPointNames:
            ws.cell(row=startIndex, column=2+(3*pointIndex)).value = returnedPointName
            ws.cell(row=startIndex, column=2+(3*pointIndex+1)).value = 'Calc_' + returnedPointName
            ws.cell(row=startIndex, column=2+(3*pointIndex+2)).value = 'Error_' + returnedPointName
            pointIndex = pointIndex + 1

        for tsIndex in range(0, len(edgeTimestamps)):
            ts = edgeTimestamps[tsIndex]
            ws.cell(row=startIndex+1+tsIndex, column=1).value = ts

            for pointIndex in range(0, len(returnedPointNames)):
                valuesAtTimestamp = timestampsAndValues[ts]
                edgePointValue = valuesAtTimestamp[pointIndex]
                ws.cell(row=startIndex+1+tsIndex, column=3*pointIndex + 2).value = edgePointValue

        self.workbook.save(self.resultsFilePath)


    def getPointNameToColumnMap(self, returnedPointNames ):
        pointNameToColumnMap = {}
        pointIndex = 0
        for pointName in returnedPointNames :
            pointNameToColumnMap[ pointName ] = pointIndex
            pointIndex = pointIndex + 1

        return pointNameToColumnMap



    def calc_kWTon(self, pointNameToColumnMap):
        sheet = self.workbook.get_sheet_by_name('CalculatedPoints')

        #kwTon = totalkW / totalTon




class E3OSHelper():

    def __init__( self, configSettings, configDirPath ):

        self.e3osSqlHost = configSettings['E3OS_SQL_HOST']
        self.e3osSqlUser = configSettings['E3OS_SQL_USER']
        self.e3osSqlPassword = configSettings['E3OS_SQL_PASSWORD']

        self.configDirPath = configDirPath
        self.headers = {'content-type': 'application/json'}

    #========= GET Data from SQL ==================================
    def private_getTableString(self, qualifier, pointsAndTypes):
        index = 1
        tableString = ""

        for pointAndType in pointsAndTypes:
            if len(tableString) > 0 :
                tableString += ' union all \n '

            pointName = pointAndType['e3osPointName']

            #rowString = "select %s, '%s%s'" % ( str(index), qualifier, pointName )
            rowString = "select " + str(index) + ", '" + qualifier + pointName + "'"
            index += 1
            tableString += rowString

        return tableString

    def getDataInRows(self, qualifier, pointsAndTypes, fromDateString, toDateString):

        queryString = '''
        declare @DataPointsOfInterest fact.datapointsofinterest
            insert @DataPointsOfInterest (seqNbr, DataPointXID)

        %s

        exec fact.DataSeriesGet2 @DataPointsOfInterest = @DataPointsOfInterest
           , @FromTime_Local = '%s'
           , @ToTime_Local = '%s'
           , @TimeInterval = 'minute'
           , @UserName = 'tkitchen'
           , @IncludeOutOfBounds = false
           , @IncludeUncommissioned = false
           , @UserId = '1d355ea9-7b16-4822-9483-1dd34173b5b8'
           , @TimeRange = null
        delete @DataPointsOfInterest
        ''' % (self.private_getTableString( qualifier, pointsAndTypes) , fromDateString, toDateString)

        try:
            conn = pymssql.connect(host=self.e3osSqlHost, user=self.e3osSqlUser, password=self.e3osSqlPassword, as_dict=True, database='oemvmdata')
        except Exception, e:
            logError( "could not connect to sql: " + str(e) )
        cur = conn.cursor()
        cur.execute( queryString )

        rowCount = 0
        rows = []
        for row in cur:
            rowCount += 1
            rows.append( row )

        conn.close()

        return self.private_transformRowData( pointsAndTypes, rowCount, rows )


    #============= Transform =======================

    def private_transformRowData(self, pointsAndTypes, rowCount, rows):

        timeStampsAndValues = {}
        timeStamps = []
        numPoints = len(pointsAndTypes)

        for row in rows:

            index = row['id'] - 1
            tz = row['tz'] #TODO?: use timezone to convert to UTC
            ts = row['time']
            #ts = ts - datetime.timedelta( minutes = tz)
            ts = ts.replace(second=0,microsecond=0)

            timestamp = ts.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] +'Z'

            val = row['value']

            if not timestamp in timeStampsAndValues:
                timeStamps.append(timestamp)
                temp = []
                for i in range (0, numPoints):
                    temp.append(None)
                timeStampsAndValues[timestamp] = temp

            vals = timeStampsAndValues[timestamp]
            vals[index] = val

        return (timeStamps, timeStampsAndValues)

    def dumpTimestampsAndValues(self, pointsAndTypes, timestamps, timeStampsAndValues):
        print 'Timestamp ',
        for pointNameAndType in pointsAndTypes:
            pointName = pointNameAndType['name']
            print pointName + " ",
        print

        for timestamp in timestamps:
            print timestamp,
            vals = timeStampsAndValues[timestamp]
            for val in vals:
                print val,
                print ' ',
            print

    #===============================


class EdgeHelperClass():

    def __init__(self, envType, configSettings, siteSid, stationId, stationName, sendingStationName, lastSuccessTimestamp):

        self.siteSid = siteSid
        self.stationId = stationId
        self.stationName = stationName
        self.sendingStationName = sendingStationName
        self.lastSuccessTimestamp = lastSuccessTimestamp

        if envType == 'PROD':
            self.host = configSettings['PROV_PROD_HOST']
            self.username = configSettings['PROV_PROD_USER']
            self.password = configSettings['PROV_PROD_PASSWORD']
        elif envType == 'OEDEV':
            self.host = configSettings['PROV_OEDEV_HOST']
            self.username = configSettings['PROV_OEDEV_USER']
            self.password = configSettings['PROV_OEDEV_PASSWORD']
        elif envType == 'OMNIBUS':
            self.host = configSettings['PROV_OMNIBUS_HOST']
            self.username = configSettings['PROV_OMNIBUS_USER']
            self.password = configSettings['PROV_OMNIBUS_PASSWORD']
        else:
            self.fatalError('no such type: ' + envType, None)

        url = self.host + "/auth/oauth/token"

        headers = {}
        headers['content-type'] = 'application/x-www-form-urlencoded'

        dict = {}
        dict['grant_type'] = 'password'
        dict['username'] = self.username
        dict['password'] = self.password
        dict['scope'] = 'read+write'

        payload = urllib.urlencode(dict)

        try:
            resp = requests.post(url, data=payload, headers=headers)
            dict = json.loads(resp.text)

        except Exception, e:
            print e

        if resp.status_code != 200:
            msg = "Can't get token! host= " + self.host
            self.fatalError(msg, resp)

        self.tok = dict['access_token']

    def getPointsList(self, sid):

        headers = {}
        headers['Accept'] = 'application/json'
        headers['content-type'] = 'application/json'
        headers['Authorization'] = 'Bearer ' + self.tok

        url = self.host + '/datapoints?sid=%s' % sid

        resp = requests.get(url, headers=headers)

        if resp.status_code != 200:
            print "Could not get points"

        dict = json.loads(resp.text)

        return dict

    def getConfigPoints(self, siteSid ):

        configPointNames = ['TotalCapacity', 'MinimumChilledWaterFlow', 'BlendedUtilityRate', 'CO2EmissionFactor']

        listOfPoints = self.getPointsList( siteSid )

        configPoints = {}
        for listEntry in listOfPoints:
            pointName = listEntry['name']

            if pointName in configPointNames:
                configPoints[ pointName ] = listEntry['value']

        return configPoints


    def allPointsAreValid(self, edisonPointsList, pointsAndTypes ):

        allPointsAreValid = True
        for pointAndTypeName in pointsAndTypes:
            pointName = pointAndTypeName['edisonPointName']

            foundThisPoint = False
            for ep in edisonPointsList:
                if ep['name'] == pointName:
                    foundThisPoint = True

            if not foundThisPoint:
                print 'could not find this point: ' + pointName
                allPointsAreValid = False

        if not allPointsAreValid:
            print 'some points invalid'
        return allPointsAreValid

    def private_getHistoryQueryParams(self, sid, pointsAndTypes, startDateString, endDateString, resolution, sparseFlag):

        pointNamesDecoratedWithAggregation = []

        for pointAndTypeName in pointsAndTypes:

            pointName = pointAndTypeName['edisonPointName']

            aggType = pointAndTypeName['aggregateTypeName']


            # cleanName = urllib.urlencode( pointName, False)
            cleanName = pointName.encode('utf-8')

            if aggType == 'None':
                pointNamesDecoratedWithAggregation.append(cleanName)
            else:
                decoratedName = aggType + "(" + cleanName + ")";
                pointNamesDecoratedWithAggregation.append(decoratedName)


        paramPoints = []
        for decortatedPoint in pointNamesDecoratedWithAggregation :
            paramPoints.append("names=" + decortatedPoint);
        pointsListAsString = '&'.join(paramPoints);

        queryString = 'sid=%s&startDate=%s&endDate=%s&resolution=%s&sparse=%s&%s' % (
            sid,
            startDateString,
            endDateString,
            resolution,
            sparseFlag,
            pointsListAsString )

        return queryString;

    def getHistory(self, sid, pointsAndTypes, startDateString, endDateString, resolution, sparseFlag ):

        headers = {}
        headers['Accept'] = 'application/json'
        headers['content-type'] = 'application/json'
        headers['Authorization'] = 'Bearer ' + self.tok

        url = self.host + '/datapoint-histories?' + self.private_getHistoryQueryParams(
            sid,
            pointsAndTypes,
            startDateString,
            endDateString,
            resolution,
            sparseFlag)


        resp = requests.get(url, headers=headers)
        if resp.status_code != 200:
             print "Could not get history"

        listOfPoints = json.loads(resp.text)

        nullValues = []
        for idx in range( 0, len(listOfPoints)):
            nullValues.append( None )

        edgeTimestamps = []
        edgeTimestampsAndValues = {}
        pointIndex = 0
        returnedPointNames = []

        for pointDict in listOfPoints:

            returnedPointNames.append( pointDict['name'])

            #this assumes each point has the same timestamps
            if len(edgeTimestamps) <= 0:
                edgeTimestamps = pointDict['timestamps']

            timestampsReturned = pointDict['timestamps']
            valuesReturned = pointDict['values']

            tsIndex = 0
            for ts in timestampsReturned:
                if not edgeTimestampsAndValues.has_key( ts ):
                    edgeTimestampsAndValues[ts] = []
                    for idx in range(0, len(listOfPoints)):
                        edgeTimestampsAndValues[ts].append(None)

                edgeValuesAtTimestamp = edgeTimestampsAndValues.get(ts)

                #hack
                if tsIndex < len( valuesReturned ):
                    edgeTimestampsAndValues[ts][pointIndex] = valuesReturned[tsIndex]
                else:
                    edgeTimestampsAndValues[ts][pointIndex] = '?'

                tsIndex = tsIndex + 1

            pointIndex = pointIndex + 1

        sss = json.dumps( listOfPoints )

        return (returnedPointNames, edgeTimestamps, edgeTimestampsAndValues )


    def fatalError(self, msg, resp):
        try:
            print msg, resp.status_code
            print str(resp.json())
        except:
            print msg
            pass
        exit()

    def getRawPoints_kW(self, stationSid, edisonPoints ):

        for pointDict in edisonPoints :
            pointName = pointDict['name']
            uom = pointDict['unitOfMeasure']
            sid = pointDict['sid']


#=================================================================
def logError( msg = 'error' ):
    print msg
    print 'example usage: python CloneUsers.py PROD OEDEV'
    exit()


if __name__ == '__main__':

    lastSuccessTimestamp = datetime.datetime.now()
    lastSuccessTimestampString = lastSuccessTimestamp.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] +'Z'

    configHelper = ConfigurationHelper()
    configSettings = configHelper.getConfigSettings()


    #========== Blue Water ===========
    #fromDateString = '2017-07-01T00:00:00.000Z'
    #toDateString = '2017-07-02T23:59:55.000Z'
    #envType = 'OEDEV'
    #stationSid = 'c:jacksonstreetindustries.s:bluewater-edge.st:1'
    #siteSid = 'c:jacksonstreetindustries.s:bluewater-edge'
    #stationID = 'bluewater_19fa9ff5'
    #stationName = 'bluewater_19fa9ff5'
    #sendingStationName = 'bluewater_19fa9ff5'
    #pointsListFileName = 'bluewater_points.xlsx'
    #qualifier = 'THPH.THC.THCEDGE.THCEDGE.'

    #========== DEMO on OEDEV ===========
    #fromDateString = '2017-07-01T00:00:00.000Z'
    #toDateString = '2017-08-22T23:59:55.000Z'
    #envType = 'OEDEV'
    #stationSid = 'c:testcustomerdemo.s:demo-edge.st:1'
    #siteSid = 'c:testcustomerdemo.s:demo-edge'
    #stationID = 'demo_f0258b66'
    #stationName = 'demo_f0258b66'
    #sendingStationName = 'demo_f0258b66'
    #pointsListFileName = 'livepoints.xlsx'
    #qualifier = 'THPH.THC.THCEDGE.THCEDGE.'

    #============ DEMO2 on PROD =======================
    #fromDateString = '2017-09-01T00:00:00.000Z'
    #toDateString = '2017-09-02T00:00:00.000Z'
    #envType = 'PROD'
    #stationSid = 'c:testcustomerdemo.s:demo2-edge.st:1'
    #siteSid = 'c:testcustomerdemo.s:demo2-edge'
    #stationID = 'demo2_358025ae'
    #stationName = 'demo2_358025ae'
    #sendingStationName = 'demo2_358025ae'
    #reportDataFileName = 'reportData.xlsx'
    #qualifier = 'JnJ.EthiconNM.EthiconNM.EthiconNM.'
    #========================================



    #============ Atom test =======================
    #fromDateString = '2017-01-01T00:00:00.000Z'
    #toDateString = '2017-01-26T23:59:55.000Z'
    #envType = 'OMNIBUS'
    #stationSid = 'c:greenstreetindustries.s:greentree-edge.st:1'
    #siteSid = 'c:greenstreetindustries.s:greentree-edge'
    #stationID = 'greentree_803b03de'
    #stationName = 'greentree_803b03de'
    #sendingStationName = 'greentree_803b03de'
    #pointsListFileName = 'VistakonAtomPoints.xlsx'
    #qualifier = 'THPH.THC.THCEDGE.THCEDGE.'
    #========================================


    print 'Opening spreadsheet...'
    xlsHelper = ExcelHelper( configHelper.getConfigDirPath(), reportDataFileName)

    print 'Getting the pointslist from excel...'
    pointsAndTypes = xlsHelper.getExcelPointsAndTypes('Points')

    print 'Connecting to sql...'
    e3oshelper = E3OSHelper( configSettings, configHelper.getConfigDirPath() )

    print 'Getting auth token...'
    edgeHelper = EdgeHelperClass(envType, configSettings, stationSid, stationID, stationName, sendingStationName, lastSuccessTimestampString)

    print 'Getting e3os history...'
    e3osTimestamps, e3osTimestampsAndValues = e3oshelper.getDataInRows( qualifier, pointsAndTypes, fromDateString, toDateString )

    # get the pointsNames from edison to make sure everything is valid
    print 'getting the station points list...'
    edisonPointsList = edgeHelper.getPointsList( stationSid )

    if not edgeHelper.allPointsAreValid(edisonPointsList, pointsAndTypes):
        print 'Some points are invalid'
        exit()

    print 'Getting edgeHistory...'
    returnedPointNames, edgeTimestamps, edgeTimestampsAndValues = edgeHelper.getHistory(stationSid, pointsAndTypes, fromDateString, toDateString, 'fiveMinutes', 'True' )

    print 'Merging History Results...'
    xlsHelper.mergeResults( returnedPointNames, e3osTimestamps, e3osTimestampsAndValues, edgeTimestamps, edgeTimestampsAndValues )

    print 'Getting the dependent points list'
    pointsFromExcel = xlsHelper.getExcelPointsAndTypes('DependentPoints')

    print 'Getting Dependent Point History...'
    returnedPointNames, edgeTimestamps, timestampsAndValues = edgeHelper.getHistory(siteSid, pointsFromExcel, fromDateString, toDateString, 'fiveMinutes', 'True' )

    print 'Getting Config Points...'
    configPoints = edgeHelper.getConfigPoints( siteSid )

    print 'Adding dependent point History Sheet'
    xlsHelper.addDependentPointHistorySheet( configPoints, returnedPointNames, edgeTimestamps, timestampsAndValues )

    print 'Getting the cacluated points list'
    pointsFromExcel = xlsHelper.getExcelPointsAndTypes('CalculatedPoints')

    print 'Get Calculated Points History'
    returnedPointNames, edgeTimestamps, timestampsAndValues = edgeHelper.getHistory(siteSid, pointsFromExcel, fromDateString, toDateString, 'fiveMinutes', 'True' )

    print 'Adding calcuated point History Sheet'
    xlsHelper.addCalculatedPointHistorySheet( configPoints, returnedPointNames, edgeTimestamps, timestampsAndValues )

    print 'Done.'

